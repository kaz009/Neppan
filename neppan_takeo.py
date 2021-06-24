# -*- coding: utf-8 -*-
import openpyxl
import datetime
import jpholiday
from selenium import webdriver
from selenium.webdriver.common.action_chains import ActionChains
from selenium.webdriver.common.alert import Alert
import time

def input_value(values):
    RType = [6,7,8,9,10,12,13,14,15,16,18,19,20,21]
    days = [i for i in range(3,17)]
    
   
    for i in range(len(RType)):
        for j in range(len(days)):
            try:
                value = values[i][j]
                XPath = "/html/body/div/table/tbody/tr/td/div[2]/table[1]/tbody/tr[1]/td/form[2]/table[4]/tbody/tr/td/table/tbody/tr/td/table/tbody/tr[" + str(RType[i]) + "]/td[" + str(days[j]) + "]/table/tbody/tr[2]/td[1]/input"
                inputField = driver.find_element_by_xpath(XPath)
                inputField.clear()
                inputField.send_keys(value)
            except:
                pass



wb = openpyxl.load_workbook("ねっぱん.xlsx", data_only=True)
ws = wb.worksheets[0]

wb_setting = openpyxl.load_workbook("設定.xlsx", data_only=True)
ws_setting = wb_setting.worksheets[0]
sgl= ws_setting["C3"].value
sd = ws_setting["C4"].value
sgl_holi=ws_setting["E3"].value
sd_holi=ws_setting["E4"].value

row =2
clm =0




for i in range(4) :
    today = datetime.date.today()
    for room in list(ws.rows)[row]:        
              
        year = today.year
        month = today.month
        
        weekday = today.weekday() 
        
        today += datetime.timedelta(days=1) 
        day = today.day
        holiday = jpholiday.is_holiday(datetime.date(year,month,day))
        
        room = int(room.value)
        if weekday == 5 or holiday == True :
            if row == 2 or row == 4 :
                room = (room * sgl_holi) // 10                
            elif row == 3 or row == 5 :
                room = (room * sd_holi) // 10                
        else :
            if row == 2 or row == 4 :
                room = (room * sgl) // 10               
            elif row == 3 or row == 5 :
                room = (room * sd) // 10
        ws.cell(column = clm+1, row = row+1, value = room)
        clm +=1
        
    clm=0
    row += 1
    
#wb.save('ねっぱん.xlsx')

values = [[ cell.value for cell in list(ws.rows)[i] ]for i in range(2,ws.max_row)]

#ねっぱんログイン
driver = webdriver.Chrome()
driver.implicitly_wait(30)

loginUrl= "https://www10.neppan.net/login.php"
driver.get(loginUrl)

username = "B01L3757"
userid = "RPAtakeo"
password = "takeo1234"

userNameField = driver.find_element_by_xpath(
    "/html/body/div/table/tbody/tr/td/div[2]/table/tbody/tr/td/form/div/table[1]/tbody/tr[2]/td/table/tbody/tr/td/table/tbody/tr[3]/td[3]/input"
    )
userNameField.send_keys(username)

userIdField = driver.find_element_by_xpath(
    "/html/body/div/table/tbody/tr/td/div[2]/table/tbody/tr/td/form/div/table[1]/tbody/tr[2]/td/table/tbody/tr/td/table/tbody/tr[4]/td[3]/input"
    )
userIdField.send_keys(userid)

passwordField = driver.find_element_by_xpath(
    "/html/body/div/table/tbody/tr/td/div[2]/table/tbody/tr/td/form/div/table[1]/tbody/tr[2]/td/table/tbody/tr/td/table/tbody/tr[5]/td[3]/input"
    )
passwordField.send_keys(password)

loginButton = driver.find_element_by_xpath(
    "/html/body/div/table/tbody/tr/td/div[2]/table/tbody/tr/td/form/div/table[1]/tbody/tr[2]/td/table/tbody/tr/td/table/tbody/tr[7]/td[3]/a"
    )
loginButton.click()

position = driver.find_element_by_xpath(
    "/html/body/div[1]/table/tbody/tr/td/div[1]/table/tbody/tr/td/table[2]/tbody/tr/td[2]/table/tbody/tr[2]/td[1]/table/tbody/tr/td/div/ul/li[2]"
    )
actions = ActionChains(driver)
actions.move_to_element(position)
actions.perform()

button = driver.find_element_by_xpath(
    "/html/body/div[1]/table/tbody/tr/td/div[1]/table/tbody/tr/td/table[2]/tbody/tr/td[2]/table/tbody/tr[2]/td[1]/table/tbody/tr/td/div/ul/li[2]/ul/li[2]/a"
    )
button.click()


#入力開始
while len(values[0]) > 14:
    input_value(values)
    for i in range(len(values)):
        del(values[i][:14])
    
    button = driver.find_element_by_xpath(
        "/html/body/div/table/tbody/tr/td/div[2]/table[1]/tbody/tr[1]/td/form[2]/div[5]/a"
        )
    actions = ActionChains(driver)
    actions.move_to_element(button)
    actions.perform()
    button.click()
    
    Alert(driver).accept()
    time.sleep(5)
    
    button = driver.find_element_by_xpath(
        "/html/body/div/table/tbody/tr/td/div[2]/table[1]/tbody/tr[1]/td/form[2]/div[4]/table/tbody/tr/td/table/tbody/tr/td/table/tbody/tr/td[7]/a"
        )
    actions = ActionChains(driver)
    actions.move_to_element(button)
    actions.perform()
    button.click()
    
   
    
        
input_value(values)
button = driver.find_element_by_xpath(
    "/html/body/div/table/tbody/tr/td/div[2]/table[1]/tbody/tr[1]/td/form[2]/div[5]/a"
    )
actions = ActionChains(driver)
actions.move_to_element(button)
actions.perform()
button.click()
Alert(driver).accept()

driver.close()